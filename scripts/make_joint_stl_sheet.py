"""Build an Excel workbook cross-referencing every STL/OBJ mesh with the joint
that moves it, so the CAD -> MJCF -> hardware chain can be verified by eye.

Nothing here is hand-typed: the MJCF is the source for joint axes/limits and the
body tree, the URDF for masses and parent/child links, calibration.json for the
per-servo home/sign, and the STL files themselves for triangle counts and
bounding boxes. Re-run after any CAD re-export.

    python -m scripts.make_joint_stl_sheet
"""

from __future__ import annotations

import json
import math
import struct
import sys
import xml.etree.ElementTree as ET
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

from hardware.k2_conventions import (  # noqa: E402
    JOINT_TO_ID,
    JOINT_TO_MJCF,
    LIMITS,
    POS_DESC,
    SIM_ORDER,
    load_calibration,
)

MJCF = ROOT / "mjcf" / "k2_physics.xml"
URDF = ROOT / "urdf" / "K2_fusion_export.urdf"
MESHDIR = ROOT / "meshes" / "K2"
OUT = ROOT / "docs" / "K2_joint_stl_map.xlsx"

HDR_FILL = PatternFill("solid", fgColor="1F3864")
HDR_FONT = Font(color="FFFFFF", bold=True, size=10)
R_FILL = PatternFill("solid", fgColor="FCE4E4")   # right-leg rows
L_FILL = PatternFill("solid", fgColor="E4EEFC")   # left-leg rows
WARN_FILL = PatternFill("solid", fgColor="FFF2CC")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


# --- parsing ---------------------------------------------------------------

def parse_mjcf():
    """joint name -> {body, axis, range, meshes} plus mesh name -> file."""
    root = ET.parse(MJCF).getroot()
    meshes = {m.get("name"): m.get("file")
              for m in root.iter("mesh") if m.get("name")}

    joints, tree, masses = {}, {}, {}

    def walk(body, parent):
        name = body.get("name")
        tree[name] = parent
        geoms = [g.get("mesh") for g in body.findall("geom")
                 if g.get("type") == "mesh"]
        inertial = body.find("inertial")
        if inertial is not None and inertial.get("mass"):
            masses[name.lower()] = float(inertial.get("mass"))
        for j in body.findall("joint"):
            joints[j.get("name")] = {
                "body": name,
                "parent_body": parent,
                "axis": j.get("axis"),
                "range": j.get("range"),
                "frcrange": j.get("actuatorfrcrange"),
                "armature": j.get("armature"),
                "meshes": geoms,
            }
        for child in body.findall("body"):
            walk(child, name)

    for body in root.find("worldbody").findall("body"):
        walk(body, "world")
    return joints, meshes, tree, masses


def parse_urdf():
    """link -> {mass, visual, collision}; and joints keyed by child link."""
    root = ET.parse(URDF).getroot()

    links = {}
    for link in root.findall("link"):
        entry = {"mass": None, "visual": None, "collision": None,
                 "com": None, "inertia": {}}
        inertial = link.find("inertial")
        if inertial is not None:
            m = inertial.find("mass")
            if m is not None:
                entry["mass"] = float(m.get("value"))
            o = inertial.find("origin")
            if o is not None:
                entry["com"] = o.get("xyz")
            i = inertial.find("inertia")
            if i is not None:
                entry["inertia"] = dict(i.attrib)
        for tag in ("visual", "collision"):
            el = link.find(tag)
            if el is not None:
                mesh = el.find("geometry/mesh")
                if mesh is not None:
                    entry[tag] = Path(mesh.get("filename")).name
        links[link.get("name")] = entry

    # Key by child link: the MJCF and URDF disagree on some joint NAMES
    # (urdf `anke_roll_ankle_pitch_*` vs mjcf `ankle_roll_foot_*`), but the
    # child link they drive is unambiguous.
    joints_by_child, joints_by_name = {}, {}
    for j in root.findall("joint"):
        entry = {
            "name": j.get("name"),
            "type": j.get("type"),
            "parent": j.find("parent").get("link"),
            "child": j.find("child").get("link"),
            "axis": j.find("axis").get("xyz") if j.find("axis") is not None else "",
            "origin": j.find("origin").get("xyz") if j.find("origin") is not None else "",
            "rpy": j.find("origin").get("rpy") if j.find("origin") is not None else "",
        }
        lim = j.find("limit")
        if lim is not None:
            entry["lower"] = float(lim.get("lower"))
            entry["upper"] = float(lim.get("upper"))
        joints_by_child[entry["child"]] = entry
        joints_by_name[entry["name"]] = entry
    return links, joints_by_child, joints_by_name


# The MJCF body names are lowercase (hip_pitch_right) where the URDF link names
# are capitalised (Hip_pitch_right), and only Knee_*/Feet_* happen to agree.
# Match on the casefolded name so every row resolves.
def ci(d: dict) -> dict:
    return {k.lower(): v for k, v in d.items()}


def stl_stats(path: Path, scale: float = 0.01):
    """Triangle count and axis-aligned bounding box (mm) of a binary STL."""
    data = path.read_bytes()
    if data[:5] == b"solid" and b"facet normal" in data[:512]:
        return {"triangles": data.count(b"facet normal"), "bbox": None,
                "bytes": len(data), "format": "ascii"}
    n = struct.unpack("<I", data[80:84])[0]
    lo = [math.inf] * 3
    hi = [-math.inf] * 3
    off = 84
    for _ in range(n):
        for v in range(3):
            base = off + 12 + v * 12
            xyz = struct.unpack("<3f", data[base:base + 12])
            for k in range(3):
                lo[k] = min(lo[k], xyz[k])
                hi[k] = max(hi[k], xyz[k])
        off += 50
    bbox = tuple((hi[k] - lo[k]) * scale * 1000.0 for k in range(3))  # -> mm
    return {"triangles": n, "bbox": bbox, "bytes": len(data),
            "format": "binary", "min": lo, "max": hi}


# --- sheet helpers ---------------------------------------------------------

def write_header(ws, headers, widths):
    ws.append(headers)
    for c, (h, w) in enumerate(zip(headers, widths), start=1):
        cell = ws.cell(row=1, column=c)
        cell.fill, cell.font, cell.border = HDR_FILL, HDR_FONT, BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.row_dimensions[1].height = 32
    ws.freeze_panes = "A2"


def style_rows(ws, ncols, side_of_row):
    for r in range(2, ws.max_row + 1):
        fill = side_of_row(r)
        for c in range(1, ncols + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=False)
            cell.font = Font(size=10)
            if fill:
                cell.fill = fill


def deg(x):
    return round(math.degrees(x), 2)


# --- sheets ----------------------------------------------------------------

def sheet_joints(wb, mj_joints, mj_meshes, u_links, u_by_child, calib, mj_masses):
    ws = wb.create_sheet("Joint ↔ STL")
    write_header(ws, [
        "Obs\nidx", "Short name", "Servo\nID", "Side", "MJCF joint name",
        "URDF joint name", "Parent link", "Child link (moved by this joint)",
        "Child collision STL", "Child visual OBJ",
        "MJCF mass (kg)\n[authoritative]", "URDF mass (kg)\n[stale]",
        "MJCF axis", "URDF axis", "Limit lo (deg)", "Limit hi (deg)",
        "home_raw", "sign", "Positive direction moves…", "OK?",
    ], [5, 15, 6, 6, 34, 34, 16, 30, 30, 26, 14, 13, 11, 14, 11, 11, 9, 6, 52, 7])

    u_by_child, u_links = ci(u_by_child), ci(u_links)
    for i, short in enumerate(SIM_ORDER):
        mjname = JOINT_TO_MJCF[short]
        mj = mj_joints[mjname]
        child = mj["body"]
        u = u_by_child.get(child.lower(), {})
        link = u_links.get(child.lower(), {})
        lo, hi = LIMITS[short]
        # The MJCF geom list is authoritative for which mesh this body draws.
        col = next((m for m in mj["meshes"] if m.endswith("_collision")), None)
        vis = next((m for m in mj["meshes"] if not m.endswith("_collision")), None)
        ws.append([
            i, short, JOINT_TO_ID[short], "R" if short.endswith("_R") else "L",
            mjname, u.get("name", "—"), u.get("parent", mj["parent_body"]), child,
            mj_meshes.get(col, "—"), mj_meshes.get(vis, "—"),
            mj_masses.get(child.lower(), "—"),
            round(link["mass"], 6) if link.get("mass") is not None else "—",
            mj["axis"], u.get("axis", "—"), deg(lo), deg(hi),
            calib["home_raw"][short], calib["sign"][short],
            POS_DESC[short], "",
        ])

    style_rows(ws, 20, lambda r: R_FILL if ws.cell(r, 4).value == "R" else L_FILL)
    for r in range(2, ws.max_row + 1):
        ws.cell(r, 20).fill = PatternFill("solid", fgColor="FFFFFF")
        ws.cell(r, 12).font = Font(size=10, italic=True, color="A6A6A6")
    return ws


def sheet_meshes(wb, mj_meshes, u_links, mj_joints, mj_masses):
    ws = wb.create_sheet("Meshes")
    write_header(ws, [
        "Link / body", "Driven by joint (short)", "Collision STL", "STL size (KB)",
        "Triangles", "BBox X (mm)", "BBox Y (mm)", "BBox Z (mm)",
        "Visual OBJ", "MJCF mass (kg)", "URDF mass (kg)", "CoM xyz (m)",
    ], [20, 22, 32, 13, 11, 12, 12, 12, 26, 14, 14, 30])

    driver = {}
    for short, mjname in JOINT_TO_MJCF.items():
        driver[mj_joints[mjname]["body"].lower()] = short

    total_u, total_mj = 0.0, 0.0
    for link, info in u_links.items():
        stl = info.get("collision")
        stats = stl_stats(MESHDIR / stl) if stl and (MESHDIR / stl).exists() else {}
        bbox = stats.get("bbox") or ("—", "—", "—")
        mass = info.get("mass")
        mj_mass = mj_masses.get(link.lower())
        total_u += mass or 0.0
        total_mj += mj_mass or 0.0
        ws.append([
            link, driver.get(link.lower(), "— (floating base)"), stl or "—",
            round(stats.get("bytes", 0) / 1024, 1) if stats else "—",
            stats.get("triangles", "—"),
            *(round(b, 2) if isinstance(b, float) else b for b in bbox),
            info.get("visual") or "—",
            mj_mass if mj_mass is not None else "—",
            round(mass, 6) if mass is not None else "—",
            info.get("com") or "—",
        ])

    ws.append(["TOTAL", "", "", "", "", "", "", "", "",
               round(total_mj, 4), round(total_u, 4), ""])
    style_rows(ws, 12, lambda r: (R_FILL if "right" in str(ws.cell(r, 1).value).lower()
                                  else L_FILL if "left" in str(ws.cell(r, 1).value).lower()
                                  else None))
    for c in range(1, 13):
        ws.cell(ws.max_row, c).font = Font(size=10, bold=True)
    for r in range(2, ws.max_row + 1):
        ws.cell(r, 11).font = Font(size=10, italic=True, color="A6A6A6")
    return ws


def sheet_notes(wb, u_by_name):
    ws = wb.create_sheet("Notes & discrepancies")
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 110
    write_header(ws, ["Topic", "What to watch for when verifying"], [22, 110])

    rows = [
        ("Sources",
         "Joint axes/limits + body tree: mjcf/k2_physics.xml. Masses, parent/child, URDF joint "
         "names: urdf/K2_fusion_export.urdf. Servo IDs, home_raw, sign: hardware/k2_conventions.py "
         "+ hardware/calibration.json. Triangle counts and bounding boxes read from the STL files."),
        ("Mesh scale",
         "Every mesh is referenced with scale=\"0.01 0.01 0.01\". The STLs are in centimetre units; "
         "bounding boxes in the Meshes sheet already have that scale applied and are shown in mm."),
        ("MASS: URDF is stale",
         "The URDF and MJCF masses disagree by about 5.6x on every link (URDF total 6.598 kg vs MJCF "
         "total 1.179 kg). The URDF is a raw Fusion export with a default material density applied to "
         "solid bodies; the MJCF carries the real measured/estimated masses and is what sim and the "
         "policies actually use. Verify against the MJCF column — the URDF column is shown greyed out "
         "only so the gap is visible. Do not regenerate the MJCF inertials from this URDF."),
        ("Joint-name mismatch",
         "The ankle-roll joint is called ankle_roll_foot_{left,right}_joint in the MJCF but "
         "anke_roll_ankle_pitch_{left,right}_joint in the URDF (note the 'anke' typo from the Fusion "
         "export). Rows are matched by CHILD LINK, not by name, so the table is still correct."),
        ("Limit mismatch",
         "joint_limits.json caps both knees at 0.0 upper even though the MJCF/CAD allows +0.2618 rad "
         "(15 deg past straight). That is deliberate — locomotion must not command the hyperextended "
         "branch. The URDF's own limits also differ from the MJCF on several joints; the MJCF is "
         "authoritative for sim and hardware."),
        ("Left/right naming",
         "The MJCF's right_* bodies sit at +y, which is the robot's own LEFT. The legs were swapped "
         "in JOINT_TO_ID on 2026-07-28 (proven on hardware): MJCF _R joints drive servos 6..11 and _L "
         "drives 0..5. Do not 'fix' the servo map from mesh names alone."),
        ("Sign checks are mirror-blind",
         "Every POS_DESC phrase is mirror-symmetric ('outward', 'pigeon-toed', 'toe up'), so a held-air "
         "check confirms each joint's axis and sign but can never catch a left/right swap. Verify side "
         "assignment by moving ONE servo ID at a time and watching which physical leg responds."),
        ("Chain order",
         "Servo IDs form a daisy-chain: one foot (0) up to that hip (5), across to the other hip (6), "
         "down to that foot (11). If a rewire changes the chain, this sheet must be regenerated."),
        ("Regenerate",
         "python -m scripts.make_joint_stl_sheet   (from the repo root)"),
    ]
    for topic, text in rows:
        ws.append([topic, text])
    for r in range(2, ws.max_row + 1):
        for c in (1, 2):
            cell = ws.cell(r, c)
            cell.border = BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.font = Font(size=10, bold=(c == 1))
            cell.fill = WARN_FILL
        ws.row_dimensions[r].height = 46
    return ws


def main():
    mj_joints, mj_meshes, _tree, mj_masses = parse_mjcf()
    u_links, u_by_child, u_by_name = parse_urdf()
    calib = load_calibration() or {"home_raw": {j: 2048 for j in SIM_ORDER},
                                   "sign": {j: 1 for j in SIM_ORDER}}

    wb = Workbook()
    wb.remove(wb.active)
    sheet_joints(wb, mj_joints, mj_meshes, u_links, u_by_child, calib, mj_masses)
    sheet_meshes(wb, mj_meshes, u_links, mj_joints, mj_masses)
    sheet_notes(wb, u_by_name)

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"wrote {OUT}")

    # Cheap consistency check so a bad export is loud rather than silent.
    missing = [j for j in SIM_ORDER if JOINT_TO_MJCF[j] not in mj_joints]
    if missing:
        print(f"WARNING: MJCF is missing joints for {missing}")
    driven = {mj_joints[JOINT_TO_MJCF[j]]["body"].lower() for j in SIM_ORDER}
    unmapped = [k for k in u_links if k.lower() not in driven]
    print(f"{len(SIM_ORDER)} joints mapped; links with no joint: {sorted(unmapped)}")


if __name__ == "__main__":
    main()
